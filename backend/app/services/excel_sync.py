import os
import glob
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings, DATA_DIR, EXPORT_DIR
from app.services.db_manager import DatabaseManager
from app.services.excel_exporter import (
    HEADER_FILL, SECTION_FILL, SUBSECTION_FILL, ZEBRA_FILL, INDEX_TAG_FILL,
    DIVIDEND_FILL, SPLIT_FILL, RESULTS_FILL, OTHER_FILL,
    HEADER_FONT, SECTION_FONT, SUBSECTION_FONT, BOLD_FONT, REGULAR_FONT, MUTED_FONT,
    THIN_BORDER, format_cell, auto_fit_columns, parse_nse_date
)

logger = logging.getLogger("excel_sync")

MASTER_INDICES_FILENAME = "broad_market_indices_master.xlsx"
MASTER_NIFTY50_FILENAME = "nifty50_daily_master.xlsx"

INDEX_CATEGORIES = [
    ("Broad Market", "Broad Market", "4F46E5"),
    ("Sectoral", "Sectoral", "0284C7"),
    ("Thematic", "Thematic", "D97706"),
    ("Strategy", "Strategy", "7C3AED"),
]

INDICES_HEADERS = [
    "Date", "Index Name", "Index Symbol", "Current Value", "Variation", "% Change",
    "Open", "High", "Low", "Prev Close", "P/E", "P/B", "Div Yield (%)",
    "Advances", "Declines", "Unchanged", "30D % Change", "365D % Change",
    "52W High", "52W Low", "1-Week Return (%)", "1-Month Return (%)", "1-Year Return (%)",
    "1-Week Ago Val", "1-Month Ago Val", "1-Year Ago Val"
]

NIFTY50_OVERVIEW_HEADERS = [
    "Date", "Symbol", "Company Name", "Series", "Open (₹)", "High (₹)", "Low (₹)",
    "Prev Close (₹)", "LTP (₹)", "Change (₹)", "% Change", "Volume (Shares)", "Turnover (₹ Cr)",
    "52W High (₹)", "52W Low (₹)", "30D % Change", "365D % Change", "Near 52W High (%)",
    "Near 52W Low (%)", "Free Float MCap (₹ Cr)", "Active Catalysts", "Last Update Time"
]

class MasterExcelSyncManager:
    """
    Manages generation, chronological appending, deduplication, and synchronization of Master Excel Workbooks:
    1. broad_market_indices_master.xlsx:
       - Sheets: Broad Market, Sectoral, Thematic, Strategy
       - Appends chronologically by Date (YYYY-MM-DD) ascending.
       - Strict deduplication on (Date, Index Symbol).
    2. nifty50_daily_master.xlsx:
       - Sheet 1: Nifty 50 Overview (22 columns, chronologically appended by date ascending).
       - Sheets 2..51: 50 Individual Stock sheets (TCS, INFY, RELIANCE, etc.) with daily trade metrics & corporate actions.
    """

    def __init__(self, output_dir: Optional[Union[str, Path]] = None):
        self.output_dir = Path(output_dir or EXPORT_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.master_indices_path = self.output_dir / MASTER_INDICES_FILENAME
        self.master_nifty50_path = self.output_dir / MASTER_NIFTY50_FILENAME

    def get_master_indices_path(self) -> str:
        return str(self.master_indices_path)

    def get_master_nifty50_path(self) -> str:
        return str(self.master_nifty50_path)

    async def build_or_rebuild_indices_master(self) -> str:
        """
        Builds or rebuilds broad_market_indices_master.xlsx from the complete historical database,
        ensuring all sheets are chronologically sorted ascending by Date and deduplicated.
        """
        logger.info(f"Rebuilding Master Indices Workbook at: {self.master_indices_path}")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for cat_title, cat_db_name, tab_color in INDEX_CATEGORIES:
            ws = wb.create_sheet(title=cat_title)
            ws.sheet_properties.tabColor = tab_color

            # Add Header Row
            ws.append(INDICES_HEADERS)
            for col_idx in range(1, len(INDICES_HEADERS) + 1):
                cell = ws.cell(row=1, column=col_idx)
                format_cell(
                    cell,
                    font=HEADER_FONT,
                    fill=HEADER_FILL,
                    align=Alignment(horizontal="center", vertical="center", wrap_text=True)
                )
            ws.row_dimensions[1].height = 28

            # Fetch all records for this category sorted by date asc, then pct_change desc
            records = await DatabaseManager.get_all_indices_history(category=cat_db_name)
            seen_keys: Set[Tuple[str, str]] = set()

            row_idx = 2
            for idx in records:
                key = (str(idx.date), str(idx.index_symbol or idx.index_name))
                if key in seen_keys:
                    continue
                seen_keys.add(key)

                pct_dec = (idx.pct_change / 100.0) if idx.pct_change is not None else None
                dy_dec = (idx.dy / 100.0) if idx.dy is not None else None
                p30_dec = (idx.per_change_30d / 100.0) if idx.per_change_30d is not None else None
                p365_dec = (idx.per_change_365d / 100.0) if idx.per_change_365d is not None else None

                val = idx.value or 0
                ret_1w = ((val - idx.one_week_ago_val) / idx.one_week_ago_val) if (idx.one_week_ago_val and idx.one_week_ago_val > 0 and val > 0) else None
                ret_1m = ((val - idx.one_month_ago_val) / idx.one_month_ago_val) if (idx.one_month_ago_val and idx.one_month_ago_val > 0 and val > 0) else None
                ret_1y = ((val - idx.one_year_ago_val) / idx.one_year_ago_val) if (idx.one_year_ago_val and idx.one_year_ago_val > 0 and val > 0) else None

                row_data = [
                    idx.date,
                    idx.index_name,
                    idx.index_symbol or idx.index_name,
                    idx.value,
                    idx.variation,
                    pct_dec,
                    idx.open,
                    idx.high,
                    idx.low,
                    idx.previous_close,
                    idx.pe,
                    idx.pb,
                    dy_dec,
                    idx.advances,
                    idx.declines,
                    idx.unchanged,
                    p30_dec,
                    p365_dec,
                    idx.year_high,
                    idx.year_low,
                    ret_1w,
                    ret_1m,
                    ret_1y,
                    idx.one_week_ago_val,
                    idx.one_month_ago_val,
                    idx.one_year_ago_val
                ]
                ws.append(row_data)

                # Format data row
                is_zebra = (row_idx % 2 == 1)
                row_fill = ZEBRA_FILL if is_zebra else None

                for c_i in range(1, len(row_data) + 1):
                    c = ws.cell(row=row_idx, column=c_i)
                    if c_i == 1:
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="center"), num_format="YYYY-MM-DD")
                    elif c_i in (2, 3):
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="left"))
                    elif c_i in (4, 7, 8, 9, 10, 19, 20, 24, 25, 26):
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                    elif c_i == 5:
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="+#,##0.00;-#,##0.00;0.00")
                    elif c_i in (6, 13, 17, 18, 21, 22, 23):
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                    elif c_i in (11, 12):
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="0.00")
                    elif c_i in (14, 15, 16):
                        format_cell(c, fill=row_fill, align=Alignment(horizontal="center"), num_format="#,##0")
                    else:
                        format_cell(c, fill=row_fill)

                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

            ws.freeze_panes = "A2"
            auto_fit_columns(ws, max_col_width=36)

        wb.save(str(self.master_indices_path))
        logger.info(f"Master Indices Workbook saved successfully ({self.master_indices_path}).")
        return str(self.master_indices_path)

    async def build_or_rebuild_nifty50_master(self) -> str:
        """
        Builds or rebuilds nifty50_daily_master.xlsx from the complete database:
        - Sheet 1: Nifty 50 Overview (all historical dates chronologically appended).
        - Sheets 2..51: 50 Individual Stock symbol sheets with daily quotes & corporate actions.
        """
        logger.info(f"Rebuilding Master Nifty 50 Workbook at: {self.master_nifty50_path}")
        wb = openpyxl.Workbook()

        # ========================================================
        # SHEET 1: NIFTY 50 OVERVIEW
        # ========================================================
        ws_overview = wb.active
        ws_overview.title = "Nifty 50 Overview"
        ws_overview.sheet_properties.tabColor = "0F766E"  # Teal

        ws_overview.append(NIFTY50_OVERVIEW_HEADERS)
        for col_idx in range(1, len(NIFTY50_OVERVIEW_HEADERS) + 1):
            cell = ws_overview.cell(row=1, column=col_idx)
            format_cell(
                cell,
                font=HEADER_FONT,
                fill=HEADER_FILL,
                align=Alignment(horizontal="center", vertical="center", wrap_text=True)
            )
        ws_overview.row_dimensions[1].height = 28

        all_stocks = await DatabaseManager.get_all_nifty50_history()
        actions_list = await DatabaseManager.get_corporate_actions()
        actions_by_symbol: Dict[str, List[Any]] = {}
        for a in actions_list:
            actions_by_symbol.setdefault(a.symbol, []).append(a)

        row_idx = 2
        seen_stock_keys: Set[Tuple[str, str]] = set()

        for s in all_stocks:
            key = (str(s.date), str(s.symbol))
            if key in seen_stock_keys:
                continue
            seen_stock_keys.add(key)

            turnover_cr = (s.turnover / 10000000.0) if s.turnover else None
            ffmc_cr = (s.ffmc / 10000000.0) if s.ffmc else None
            pct_dec = (s.pct_change / 100.0) if s.pct_change is not None else None
            p30_dec = (s.per_change_30d / 100.0) if s.per_change_30d is not None else None
            p365_dec = (s.per_change_365d / 100.0) if s.per_change_365d is not None else None
            wkh_dec = (s.near_wkh / 100.0) if s.near_wkh is not None else None
            wkl_dec = (s.near_wkl / 100.0) if s.near_wkl is not None else None

            # Check catalysts
            stock_cas = actions_by_symbol.get(s.symbol, [])
            cat_str = f"{len(stock_cas)} Events" if stock_cas else "None"

            row_data = [
                s.date,
                s.symbol,
                s.company_name or s.symbol,
                s.series or "EQ",
                s.open,
                s.high,
                s.low,
                s.previous_close,
                s.ltp,
                s.change,
                pct_dec,
                s.volume,
                turnover_cr,
                s.year_high,
                s.year_low,
                p30_dec,
                p365_dec,
                wkh_dec,
                wkl_dec,
                ffmc_cr,
                cat_str,
                s.last_update_time or "-"
            ]
            ws_overview.append(row_data)

            is_zebra = (row_idx % 2 == 1)
            row_fill = ZEBRA_FILL if is_zebra else None

            for c_i in range(1, len(row_data) + 1):
                c = ws_overview.cell(row=row_idx, column=c_i)
                if c_i == 1:
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="center"), num_format="YYYY-MM-DD")
                elif c_i in (2, 3):
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="left"))
                elif c_i == 4:
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="center"))
                elif c_i in (5, 6, 7, 8, 9, 14, 15):
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="₹#,##0.00")
                elif c_i == 10:
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="+₹#,##0.00;-₹#,##0.00;₹0.00")
                elif c_i in (11, 16, 17, 18, 19):
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                elif c_i == 12:
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0")
                elif c_i in (13, 20):
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="right"), num_format="₹#,##0.00")
                else:
                    format_cell(c, fill=row_fill, align=Alignment(horizontal="center"))

            ws_overview.row_dimensions[row_idx].height = 20
            row_idx += 1

        ws_overview.freeze_panes = "A2"
        auto_fit_columns(ws_overview, max_col_width=40)

        # ========================================================
        # SHEETS 2..51: INDIVIDUAL STOCK SHEETS (Chronological)
        # ========================================================
        distinct_symbols = sorted(list(set(s.symbol for s in all_stocks)))
        stock_details_all = await DatabaseManager.get_all_stock_details_history()
        details_map: Dict[Tuple[str, str], Any] = {}
        for det in stock_details_all:
            details_map[(str(det.date), str(det.symbol))] = det

        # Group stocks by symbol
        stocks_by_sym: Dict[str, List[Any]] = {}
        for s in all_stocks:
            stocks_by_sym.setdefault(s.symbol, []).append(s)

        for sym in distinct_symbols:
            safe_sheet_title = sym[:31].replace("/", "-").replace("\\", "-")
            ws_stock = wb.create_sheet(title=safe_sheet_title)
            ws_stock.sheet_properties.tabColor = "0284C7"

            # Header Demarcation
            ws_stock.merge_cells("A1:D1")
            title_cell = ws_stock.cell(1, 1, f"STOCK PROFILE & HISTORICAL QUOTES: {sym}")
            format_cell(title_cell, font=SECTION_FONT, fill=HEADER_FILL, align=Alignment(horizontal="center"))
            ws_stock.row_dimensions[1].height = 28

            cur_row = 3
            sym_records = stocks_by_sym.get(sym, [])
            sym_records.sort(key=lambda x: str(x.date))  # Chronological ascending

            # Chronological Daily Quotes Table
            ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
            sec_quotes = ws_stock.cell(cur_row, 1, "1. Historical Daily Market Quotes (Chronological Order)")
            format_cell(sec_quotes, font=SUBSECTION_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
            ws_stock.row_dimensions[cur_row].height = 22
            cur_row += 1

            quote_headers = ["Trade Date", "LTP (₹)", "Change (₹)", "% Change", "Open (₹)", "High (₹)", "Low (₹)", "Volume", "Turnover (₹ Cr)", "52W High", "52W Low"]
            for col_i, qh in enumerate(quote_headers, 1):
                c_qh = ws_stock.cell(cur_row, col_i, qh)
                format_cell(c_qh, font=BOLD_FONT, fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"), align=Alignment(horizontal="center"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

            for s_rec in sym_records:
                pct_d = (s_rec.pct_change / 100.0) if s_rec.pct_change is not None else None
                to_cr = (s_rec.turnover / 10000000.0) if s_rec.turnover else None
                q_row_data = [
                    s_rec.date,
                    s_rec.ltp,
                    s_rec.change,
                    pct_d,
                    s_rec.open,
                    s_rec.high,
                    s_rec.low,
                    s_rec.volume,
                    to_cr,
                    s_rec.year_high,
                    s_rec.year_low
                ]
                for col_i, val in enumerate(q_row_data, 1):
                    c = ws_stock.cell(cur_row, col_i, val)
                    if col_i == 1:
                        format_cell(c, align=Alignment(horizontal="center"), num_format="YYYY-MM-DD")
                    elif col_i in (2, 5, 6, 7, 10, 11):
                        format_cell(c, align=Alignment(horizontal="right"), num_format="₹#,##0.00")
                    elif col_i == 3:
                        format_cell(c, align=Alignment(horizontal="right"), num_format="+₹#,##0.00;-₹#,##0.00;₹0.00")
                    elif col_i == 4:
                        format_cell(c, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                    elif col_i == 8:
                        format_cell(c, align=Alignment(horizontal="right"), num_format="#,##0")
                    elif col_i == 9:
                        format_cell(c, align=Alignment(horizontal="right"), num_format="₹#,##0.00")
                    else:
                        format_cell(c)
                ws_stock.row_dimensions[cur_row].height = 19
                cur_row += 1

            cur_row += 1

            # Section 2: Corporate Actions & Dividends Timeline
            ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
            sec_ca = ws_stock.cell(cur_row, 1, "2. Corporate Actions & Dividends Timeline")
            format_cell(sec_ca, font=SUBSECTION_FONT, fill=SUBSECTION_FILL, align=Alignment(horizontal="left"))
            ws_stock.row_dimensions[cur_row].height = 22
            cur_row += 1

            ca_headers = ["Ex-Date", "Action Type", "Subject / Description", "Record Date"]
            for col_i, ch in enumerate(ca_headers, 1):
                c_ch = ws_stock.cell(cur_row, col_i, ch)
                format_cell(c_ch, font=BOLD_FONT, fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"), align=Alignment(horizontal="center"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

            stock_ca_all = actions_by_symbol.get(sym, [])
            stock_ca_all.sort(key=lambda x: parse_nse_date(x.ex_date), reverse=True)

            if stock_ca_all:
                for ca_item in stock_ca_all:
                    c1 = ws_stock.cell(cur_row, 1, ca_item.ex_date or "-")
                    c2 = ws_stock.cell(cur_row, 2, ca_item.action_type or "-")
                    c3 = ws_stock.cell(cur_row, 3, ca_item.subject or "-")
                    c4 = ws_stock.cell(cur_row, 4, ca_item.record_date or "-")

                    badge_fill = DIVIDEND_FILL if ca_item.action_type == "DIVIDEND" else (
                        SPLIT_FILL if ca_item.action_type in ("SPLIT", "BONUS") else OTHER_FILL
                    )
                    format_cell(c1, align=Alignment(horizontal="center"))
                    format_cell(c2, font=BOLD_FONT, fill=badge_fill, align=Alignment(horizontal="center"))
                    format_cell(c3, align=Alignment(horizontal="left"))
                    format_cell(c4, align=Alignment(horizontal="center"))
                    ws_stock.row_dimensions[cur_row].height = 20
                    cur_row += 1
            else:
                for col_i in range(1, 5):
                    c = ws_stock.cell(cur_row, col_i, "-")
                    format_cell(c, align=Alignment(horizontal="center"))
                ws_stock.row_dimensions[cur_row].height = 20
                cur_row += 1

            ws_stock.freeze_panes = "A3"
            auto_fit_columns(ws_stock, max_col_width=45)

        wb.save(str(self.master_nifty50_path))
        logger.info(f"Master Nifty 50 Workbook saved successfully ({self.master_nifty50_path}).")
        return str(self.master_nifty50_path)

    async def sync_all_masters(self) -> Tuple[str, str]:
        """Synchronizes and generates both master workbooks from the database."""
        indices_path = await self.build_or_rebuild_indices_master()
        nifty50_path = await self.build_or_rebuild_nifty50_master()
        return indices_path, nifty50_path

    async def append_daily_data(self, target_date: str) -> Tuple[str, str]:
        """
        Appends new daily data for the given date across all matching sheets in both master workbooks.
        Re-sorts chronologically ascending by Date and eliminates duplicates.
        """
        logger.info(f"Appending daily data for {target_date} into Master Workbooks...")
        return await self.sync_all_masters()

    async def ingest_historical_excel_files(self, paths_or_dir: Union[str, List[str]]) -> Dict[str, Any]:
        """
        Ingests existing daily multi-sheet Excel files (`broad_market_indices_*.xlsx` and `nifty50_daily_*.xlsx`),
        extracts data from all sheets, automatically classifies them, stores into the local SQLite database,
        and updates master workbooks.
        """
        file_list: List[str] = []
        if isinstance(paths_or_dir, str):
            if os.path.isdir(paths_or_dir):
                file_list = glob.glob(os.path.join(paths_or_dir, "*.xlsx"))
            elif os.path.isfile(paths_or_dir):
                file_list = [paths_or_dir]
            else:
                file_list = glob.glob(paths_or_dir)
        elif isinstance(paths_or_dir, list):
            file_list = paths_or_dir

        logger.info(f"Ingesting {len(file_list)} historical Excel file(s)...")
        indices_count = 0
        nifty50_count = 0
        classified_files = []

        for fpath in file_list:
            fname = os.path.basename(fpath)
            fname_lower = fname.lower()
            if "master" in fname_lower or fname.startswith("~$"):
                continue  # Skip existing master workbooks and lock files

            file_info = {
                "filename": fname,
                "classification": "Unknown",
                "records_imported": 0,
                "dates_detected": set()
            }

            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                
                # Classification 1: Indices Workbook
                has_indices_sheets = any(sheet in wb.sheetnames for sheet in ["Broad Market", "Sectoral", "Thematic", "Strategy"])
                
                # Classification 2: Nifty 50 Workbook
                has_nifty_sheet = "Nifty 50 Overview" in wb.sheetnames or any(sym in wb.sheetnames for sym in ["RELIANCE", "TCS", "INFY", "HDFCBANK"])

                if has_indices_sheets:
                    file_info["classification"] = "Broad Market Indices Master"
                    for sheet_name in ["Broad Market", "Sectoral", "Thematic", "Strategy"]:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            rows = list(ws.iter_rows(values_only=True))
                            if len(rows) <= 1:
                                continue
                            
                            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
                            
                            def get_col(candidates, default_idx=None):
                                for c in candidates:
                                    if c in headers:
                                        return headers.index(c)
                                return default_idx

                            date_idx = get_col(["date"], 0)
                            sym_idx = get_col(["index symbol", "index name", "index"], 1)
                            val_idx = get_col(["current value", "value", "close", "last price"], 3)
                            var_idx = get_col(["variation", "change"], 4)
                            pct_idx = get_col(["% change", "pct_change", "percent change"], 5)
                            open_idx = get_col(["open"], 6)
                            high_idx = get_col(["high"], 7)
                            low_idx = get_col(["low"], 8)
                            prev_idx = get_col(["previous close", "prev close"], 9)
                            y_high_idx = get_col(["52w high", "52 week high", "year high"], 10)
                            y_low_idx = get_col(["52w low", "52 week low", "year low"], 11)
                            pe_idx = get_col(["p/e", "pe", "pe ratio"], 12)
                            pb_idx = get_col(["p/b", "pb", "pb ratio"], 13)
                            dy_idx = get_col(["div yield (%)", "dividend yield", "dy"], 14)
                            adv_idx = get_col(["advances", "adv"], 15)
                            dec_idx = get_col(["declines", "dec"], 16)
                            unc_idx = get_col(["unchanged", "unc"], 17)
                            
                            batch_records = []
                            for row in rows[1:]:
                                if not row or not any(row):
                                    continue
                                d_val = str(row[date_idx]).strip() if date_idx is not None and date_idx < len(row) and row[date_idx] else None
                                if not d_val or d_val.lower() == "date":
                                    continue
                                try:
                                    d_parsed = parse_nse_date(d_val).strftime("%Y-%m-%d")
                                    if d_parsed != "0001-01-01":
                                        d_val = d_parsed
                                except Exception:
                                    pass

                                sym_val = str(row[sym_idx]).strip() if sym_idx is not None and sym_idx < len(row) and row[sym_idx] else ""
                                if not sym_val:
                                    continue

                                file_info["dates_detected"].add(d_val)

                                def parse_num(idx):
                                    if idx is not None and idx < len(row) and row[idx] is not None:
                                        s = str(row[idx]).replace(",", "").replace("₹", "").replace("%", "").strip()
                                        try:
                                            return float(s)
                                        except ValueError:
                                            return None
                                    return None

                                rec = {
                                    "date": d_val,
                                    "index_category": sheet_name,
                                    "index_name": sym_val,
                                    "index_symbol": sym_val,
                                    "value": parse_num(val_idx),
                                    "variation": parse_num(var_idx),
                                    "pct_change": parse_num(pct_idx),
                                    "open": parse_num(open_idx),
                                    "high": parse_num(high_idx),
                                    "low": parse_num(low_idx),
                                    "previous_close": parse_num(prev_idx),
                                    "year_high": parse_num(y_high_idx),
                                    "year_low": parse_num(y_low_idx),
                                    "pe": parse_num(pe_idx),
                                    "pb": parse_num(pb_idx),
                                    "dy": parse_num(dy_idx),
                                    "advances": int(parse_num(adv_idx)) if parse_num(adv_idx) is not None else None,
                                    "declines": int(parse_num(dec_idx)) if parse_num(dec_idx) is not None else None,
                                    "unchanged": int(parse_num(unc_idx)) if parse_num(unc_idx) is not None else None,
                                }
                                batch_records.append(rec)

                            if batch_records:
                                cnt = await DatabaseManager.upsert_indices_records(batch_records)
                                indices_count += cnt
                                file_info["records_imported"] += cnt

                elif has_nifty_sheet:
                    file_info["classification"] = "NIFTY 50 Daily Master"
                    if "Nifty 50 Overview" in wb.sheetnames:
                        ws = wb["Nifty 50 Overview"]
                        rows = list(ws.iter_rows(values_only=True))
                        if len(rows) > 1:
                            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
                            
                            def get_col_n50(candidates, default_idx=None):
                                for c in candidates:
                                    if c in headers:
                                        return headers.index(c)
                                return default_idx

                            date_idx = get_col_n50(["date"], 0)
                            sym_idx = get_col_n50(["symbol"], 1)
                            comp_idx = get_col_n50(["company name", "company"], 2)
                            series_idx = get_col_n50(["series"], 3)
                            open_idx = get_col_n50(["open (₹)", "open"], 4)
                            high_idx = get_col_n50(["high (₹)", "high"], 5)
                            low_idx = get_col_n50(["low (₹)", "low"], 6)
                            prev_idx = get_col_n50(["prev close (₹)", "previous close", "prev close"], 7)
                            ltp_idx = get_col_n50(["ltp (₹)", "ltp", "last price"], 8)
                            chg_idx = get_col_n50(["change (₹)", "change"], 9)
                            pct_idx = get_col_n50(["% change", "pct_change"], 10)
                            vol_idx = get_col_n50(["volume (shares)", "volume"], 11)
                            turn_idx = get_col_n50(["turnover (₹ cr)", "turnover"], 12)
                            w_high_idx = get_col_n50(["52w high (₹)", "52w high"], 13)
                            w_low_idx = get_col_n50(["52w low (₹)", "52w low"], 14)
                            pe_idx = get_col_n50(["p/e", "pe ratio"], 17)
                            pb_idx = get_col_n50(["p/b", "pb ratio"], 18)
                            deliv_idx = get_col_n50(["delivery %", "delivery_pct"], 21)

                            batch_stocks = []
                            for row in rows[1:]:
                                if not row or not any(row):
                                    continue
                                d_val = str(row[date_idx]).strip() if date_idx is not None and date_idx < len(row) and row[date_idx] else None
                                if not d_val or d_val.lower() == "date":
                                    continue
                                try:
                                    d_parsed = parse_nse_date(d_val).strftime("%Y-%m-%d")
                                    if d_parsed != "0001-01-01":
                                        d_val = d_parsed
                                except Exception:
                                    pass

                                sym_val = str(row[sym_idx]).strip() if sym_idx is not None and sym_idx < len(row) and row[sym_idx] else ""
                                if not sym_val:
                                    continue

                                file_info["dates_detected"].add(d_val)

                                def parse_num_s(idx):
                                    if idx is not None and idx < len(row) and row[idx] is not None:
                                        s = str(row[idx]).replace(",", "").replace("₹", "").replace("%", "").strip()
                                        try:
                                            return float(s)
                                        except ValueError:
                                            return None
                                    return None

                                s_rec = {
                                    "date": d_val,
                                    "symbol": sym_val,
                                    "company_name": str(row[comp_idx]).strip() if comp_idx is not None and comp_idx < len(row) and row[comp_idx] else sym_val,
                                    "series": str(row[series_idx]).strip() if series_idx is not None and series_idx < len(row) and row[series_idx] else "EQ",
                                    "open_price": parse_num_s(open_idx),
                                    "day_high": parse_num_s(high_idx),
                                    "day_low": parse_num_s(low_idx),
                                    "previous_close": parse_num_s(prev_idx),
                                    "ltp": parse_num_s(ltp_idx),
                                    "change": parse_num_s(chg_idx),
                                    "pct_change": parse_num_s(pct_idx),
                                    "volume_shares": int(parse_num_s(vol_idx)) if parse_num_s(vol_idx) is not None else None,
                                    "turnover_cr": parse_num_s(turn_idx),
                                    "week52_high": parse_num_s(w_high_idx),
                                    "week52_low": parse_num_s(w_low_idx),
                                    "pe_ratio": parse_num_s(pe_idx),
                                    "pb_ratio": parse_num_s(pb_idx),
                                    "delivery_pct": parse_num_s(deliv_idx),
                                }
                                batch_stocks.append(s_rec)

                            if batch_stocks:
                                cnt = await DatabaseManager.upsert_nifty50_records(batch_stocks)
                                nifty50_count += cnt
                                file_info["records_imported"] += cnt

                file_info["dates_detected"] = sorted(list(file_info["dates_detected"]))
                classified_files.append(file_info)

            except Exception as e:
                logger.error(f"Error parsing historical Excel file {fpath}: {e}")
                file_info["error"] = str(e)
                classified_files.append(file_info)

        # Update and rebuild Master files
        await self.sync_all_masters()

        logger.info(f"Ingestion complete: {indices_count} index records, {nifty50_count} stock records imported.")
        return {
            "success": True,
            "files_processed": len(file_list),
            "indices_imported": indices_count,
            "nifty50_imported": nifty50_count,
            "classified_files": classified_files,
            "master_indices_path": str(self.master_indices_path),
            "master_nifty50_path": str(self.master_nifty50_path)
        }

# Global Instance
master_excel_sync = MasterExcelSyncManager()
