import os
import zipfile
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

from sqlalchemy import select, asc, desc
from sqlalchemy.ext.asyncio import AsyncSession

import asyncio
import logging
from app.models import Nifty50Daily, StockDetailDaily, IndexDaily, CorporateAction
from app.database import AsyncSessionLocal
from app.config import settings
from app.services.nse_fetcher import NSEFetcher, classify_corporate_action

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")      # Dark Slate
SECTION_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")     # Slate 700
SUBSECTION_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")  # Teal 700
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
INDEX_TAG_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

DIVIDEND_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")     # Soft Green
SPLIT_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")        # Soft Purple
RESULTS_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")      # Soft Blue
OTHER_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
SUBSECTION_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
REGULAR_FONT = Font(name="Segoe UI", size=10)
MUTED_FONT = Font(name="Segoe UI", size=9, italic=True, color="64748B")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)

def parse_nse_date(d_str: Optional[str]) -> datetime:
    """Parses various NSE date formats for accurate chronological sorting."""
    if not d_str or str(d_str).strip() in ("-", "", "None"):
        return datetime.min
    clean = str(d_str).strip()
    formats = [
        "%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y",
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
        "%Y%m%d", "%d%b%Y"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(clean, fmt)
        except Exception:
            pass
    return datetime.min

def auto_fit_columns(ws, max_col_width: int = 42):
    """Automatically adjusts column widths based on contents."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if "\n" in val_str:
                val_str = max(val_str.split("\n"), key=len)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_col_width)

def format_cell(cell, font=REGULAR_FONT, fill=None, align=None, num_format=None, border=THIN_BORDER):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if num_format: cell.number_format = num_format
    if border: cell.border = border

async def build_nifty50_workbook(target_date: str, output_path: str) -> str:
    """Generates nifty50_daily_YYYY-MM-DD.xlsx with corporate actions & exhaustive metrics from NSE."""
    wb = openpyxl.Workbook()
    
    # ==========================================
    # SHEET 1: NIFTY 50 OVERVIEW
    # ==========================================
    ws_overview = wb.active
    ws_overview.title = "Nifty 50 Overview"
    ws_overview.sheet_properties.tabColor = "0F766E"  # Teal

    async with AsyncSessionLocal() as db:
        q = await db.execute(
            select(Nifty50Daily).where(Nifty50Daily.date == target_date).order_by(Nifty50Daily.pct_change.desc())
        )
        stocks: List[Nifty50Daily] = q.scalars().all()

        if not stocks:
            q_latest = await db.execute(
                select(Nifty50Daily.date).order_by(Nifty50Daily.date.desc()).limit(1)
            )
            latest_d = q_latest.scalars().first()
            if latest_d:
                target_date = latest_d
                q2 = await db.execute(
                    select(Nifty50Daily).where(Nifty50Daily.date == target_date).order_by(Nifty50Daily.pct_change.desc())
                )
                stocks = q2.scalars().all()

        q_det = await db.execute(
            select(StockDetailDaily).where(StockDetailDaily.date == target_date)
        )
        details_map = {d.symbol: d for d in q_det.scalars().all()}

        # Load corporate actions for Nifty 50 symbols
        symbols_list = [s.symbol for s in stocks]
        q_ca = await db.execute(
            select(CorporateAction).where(CorporateAction.symbol.in_(symbols_list)).order_by(asc(CorporateAction.priority_level), desc(CorporateAction.ex_date))
        )
        all_actions = q_ca.scalars().all()
        actions_by_symbol: Dict[str, List[CorporateAction]] = {}
        for a in all_actions:
            actions_by_symbol.setdefault(a.symbol, []).append(a)

        # On-demand fetch for symbols that don't have actions in DB yet
        missing_ca_symbols = [s for s in symbols_list if s not in actions_by_symbol]
        if missing_ca_symbols:
            fetcher = NSEFetcher()
            sem_exp = asyncio.Semaphore(5)

            async def fetch_missing_ca(s_sym: str):
                async with sem_exp:
                    try:
                        ca_res = await asyncio.to_thread(fetcher.fetch_stock_corporate_actions, s_sym)
                        bm_res = await asyncio.to_thread(fetcher.fetch_stock_event_calendar, s_sym)
                        seen_keys = set()
                        async with AsyncSessionLocal() as sub_db:
                            for item in (ca_res or []):
                                subj = item.get("subject")
                                ex_d = item.get("exDate") or item.get("caBroadcastDate")
                                if subj and (s_sym, subj, ex_d) not in seen_keys:
                                    seen_keys.add((s_sym, subj, ex_d))
                                    act_type, priority = classify_corporate_action(subj)
                                    q_chk = await sub_db.execute(
                                        select(CorporateAction).where(
                                            CorporateAction.symbol == s_sym,
                                            CorporateAction.subject == subj,
                                            CorporateAction.ex_date == ex_d
                                        )
                                    )
                                    existing = q_chk.scalars().first()
                                    if not existing:
                                        new_act = CorporateAction(
                                            symbol=s_sym,
                                            company_name=item.get("comp"),
                                            series=item.get("series", "EQ"),
                                            subject=subj,
                                            action_type=act_type,
                                            ex_date=ex_d,
                                            record_date=item.get("recDate"),
                                            priority_level=priority,
                                            raw_data=item
                                        )
                                        sub_db.add(new_act)
                                        actions_by_symbol.setdefault(s_sym, []).append(new_act)
                                    else:
                                        actions_by_symbol.setdefault(s_sym, []).append(existing)

                            for ev in (bm_res or []):
                                purpose = ev.get("purpose") or "Board Meeting"
                                ev_date = ev.get("date")
                                if purpose and (s_sym, purpose, ev_date) not in seen_keys:
                                    seen_keys.add((s_sym, purpose, ev_date))
                                    act_type, priority = classify_corporate_action(purpose, ev.get("bm_desc") or "")
                                    q_chk_bm = await sub_db.execute(
                                        select(CorporateAction).where(
                                            CorporateAction.symbol == s_sym,
                                            CorporateAction.subject == purpose,
                                            CorporateAction.ex_date == ev_date
                                        )
                                    )
                                    existing_bm = q_chk_bm.scalars().first()
                                    if not existing_bm:
                                        new_bm = CorporateAction(
                                            symbol=s_sym,
                                            company_name=ev.get("company"),
                                            series="EQ",
                                            subject=purpose,
                                            action_type=act_type,
                                            ex_date=ev_date,
                                            details=ev.get("bm_desc"),
                                            priority_level=priority,
                                            raw_data=ev
                                        )
                                        sub_db.add(new_bm)
                                        actions_by_symbol.setdefault(s_sym, []).append(new_bm)
                                    else:
                                        actions_by_symbol.setdefault(s_sym, []).append(existing_bm)

                            try:
                                await sub_db.commit()
                            except Exception:
                                await sub_db.rollback()
                    except Exception as err:
                        logger.warning(f"Failed on-demand CA fetch in export for {s_sym}: {err}")

            exp_tasks = [fetch_missing_ca(s) for s in missing_ca_symbols]
            await asyncio.gather(*exp_tasks)

    # Exhaustive Overview Headers with Corporate Actions columns
    headers = [
        "Date", "Symbol", "Company Name", "Series", "LTP (₹)", "Change (₹)", "% Change",
        "Open (₹)", "Day High (₹)", "Day Low (₹)", "Prev Close (₹)",
        "Volume (Shares)", "Turnover (₹ Cr)", "Free Float MCap (₹ Cr)",
        "Upcoming Corporate Action", "Board Meeting / Results Date",
        "52W High (₹)", "52W Low (₹)", "30D % Change", "365D % Change",
        "Near 52W High (%)", "Near 52W Low (%)"
    ]

    ws_overview.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws_overview.cell(row=1, column=col_idx)
        format_cell(
            cell,
            font=HEADER_FONT,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True)
        )
    ws_overview.row_dimensions[1].height = 28

    # Overview Data Rows
    row_start = 2
    for s in stocks:
        turnover_cr = (s.turnover / 10000000.0) if s.turnover else None
        ffmc_cr = (s.ffmc / 10000000.0) if s.ffmc else None
        pct_change_dec = (s.pct_change / 100.0) if s.pct_change is not None else None
        p30_dec = (s.per_change_30d / 100.0) if s.per_change_30d is not None else None
        p365_dec = (s.per_change_365d / 100.0) if s.per_change_365d is not None else None
        near_h_dec = (s.near_wkh / 100.0) if s.near_wkh is not None else None
        near_l_dec = (s.near_wkl / 100.0) if s.near_wkl is not None else None

        # Corporate Action Summary sorted chronologically
        stock_actions = actions_by_symbol.get(s.symbol, [])
        ca_candidates = [a for a in stock_actions if a.action_type in ("DIVIDEND", "SPLIT", "BONUS", "BUYBACK", "RIGHTS")]
        ca_candidates.sort(key=lambda x: parse_nse_date(x.ex_date), reverse=True)

        bm_candidates = [a for a in stock_actions if a.action_type in ("RESULTS", "BOARD_MEETING", "AGM")]
        bm_candidates.sort(key=lambda x: parse_nse_date(x.ex_date), reverse=True)

        ca_summary = "-"
        if ca_candidates:
            top_ca = ca_candidates[0]
            ex_str = f" (Ex: {top_ca.ex_date})" if top_ca.ex_date else ""
            ca_summary = f"{top_ca.subject}{ex_str}"

        bm_summary = "-"
        if bm_candidates:
            top_bm = bm_candidates[0]
            bm_summary = f"{top_bm.ex_date} ({top_bm.subject})" if top_bm.ex_date else str(top_bm.subject)

        row_data = [
            s.date or target_date,
            s.symbol,
            s.company_name or s.symbol,
            s.series or "EQ",
            s.ltp,
            s.change,
            pct_change_dec,
            s.open,
            s.high,
            s.low,
            s.previous_close,
            s.volume,
            turnover_cr,
            ffmc_cr,
            ca_summary,
            bm_summary,
            s.year_high,
            s.year_low,
            p30_dec,
            p365_dec,
            near_h_dec,
            near_l_dec
        ]
        ws_overview.append(row_data)

    row_end = len(stocks) + 1

    # Format Overview Cells
    for r in range(row_start, row_end + 1):
        is_zebra = (r % 2 == 1)
        row_fill = ZEBRA_FILL if is_zebra else None

        format_cell(ws_overview.cell(r, 1), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="center"))
        format_cell(ws_overview.cell(r, 2), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="left"))
        format_cell(ws_overview.cell(r, 3), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="left"))
        format_cell(ws_overview.cell(r, 4), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="center"))
        format_cell(ws_overview.cell(r, 5), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 6), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+#,##0.00;-#,##0.00;0.00")
        format_cell(ws_overview.cell(r, 7), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
        format_cell(ws_overview.cell(r, 8), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 9), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 10), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 11), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 12), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0")
        format_cell(ws_overview.cell(r, 13), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 14), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 15), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="left"))
        format_cell(ws_overview.cell(r, 16), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="center"))
        format_cell(ws_overview.cell(r, 17), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 18), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
        format_cell(ws_overview.cell(r, 19), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
        format_cell(ws_overview.cell(r, 20), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
        format_cell(ws_overview.cell(r, 21), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
        format_cell(ws_overview.cell(r, 22), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
        ws_overview.row_dimensions[r].height = 20

    if row_end >= row_start:
        # ColorScaleRule on % Change (Col G) and 30D / 365D % Change (Col S / Col T)
        color_scale = ColorScaleRule(
            start_type="num", start_value=-0.05, start_color="FCA5A5",
            mid_type="num", mid_value=0.0, mid_color="FFFFFF",
            end_type="num", end_value=0.05, end_color="86EFAC"
        )
        ws_overview.conditional_formatting.add(f"G2:G{row_end}", color_scale)
        ws_overview.conditional_formatting.add(f"S2:T{row_end}", color_scale)

        # DataBarRule on Volume (Col L) & Turnover (Col M)
        data_bar_vol = DataBarRule(start_type="min", end_type="max", color="60A5FA", showValue="None", minLength=None, maxLength=None)
        ws_overview.conditional_formatting.add(f"L2:L{row_end}", data_bar_vol)

        data_bar_to = DataBarRule(start_type="min", end_type="max", color="93C5FD", showValue="None", minLength=None, maxLength=None)
        ws_overview.conditional_formatting.add(f"M2:M{row_end}", data_bar_to)

        # IconSetRule on % Change (Col G)
        icon_set = IconSetRule("3Arrows", "num", [0, 0.0001], showValue=None, reverse=None)
        ws_overview.conditional_formatting.add(f"G2:G{row_end}", icon_set)

    ws_overview.freeze_panes = "C2"
    ws_overview.auto_filter.ref = f"A1:V{max(row_end, 1)}"
    auto_fit_columns(ws_overview)

    # ==========================================
    # SHEETS 2 TO 51: INDIVIDUAL STOCK SHEETS
    # ==========================================
    for stock in stocks:
        sym = stock.symbol
        ws_stock = wb.create_sheet(title=sym[:31])
        ws_stock.sheet_properties.tabColor = "3B82F6"  # Blue
        detail = details_map.get(sym)

        t_info = (detail.trade_info if detail and isinstance(detail.trade_info, dict) else {})
        p_info = (detail.price_info if detail and isinstance(detail.price_info, dict) else {})
        s_info = (detail.security_info if detail and isinstance(detail.security_info, dict) else {})
        m_data = (detail.meta_data if detail and isinstance(detail.meta_data, dict) else {})

        # Header Title Banner
        ws_stock.merge_cells("A1:D1")
        title_cell = ws_stock.cell(1, 1, f"{stock.company_name or sym} ({sym}) — Complete Daily Analytics & Corporate Actions [{stock.date or target_date}]")
        format_cell(title_cell, font=SECTION_FONT, fill=HEADER_FILL, align=Alignment(horizontal="left", vertical="center"))
        ws_stock.row_dimensions[1].height = 32

        cur_row = 3

        # ----------------------------------------------------
        # SECTION 1: PRICE & VALUATION PERFORMANCE OVERVIEW
        # ----------------------------------------------------
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec1 = ws_stock.cell(cur_row, 1, "1. Price, Valuation & Performance Overview")
        format_cell(sec1, font=HEADER_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 24
        cur_row += 1

        vwap_val = p_info.get("vwap") or m_data.get("averagePrice")
        yr_high_dt = str(p_info.get("yearHightDt") or "-")
        yr_low_dt = str(p_info.get("yearLowDt") or "-")
        price_band = str(p_info.get("priceBand") or "No Band")
        tick_size = str(p_info.get("tickSize") or "0.05")

        price_rows = [
            ("Date", str(stock.date or target_date), "Symbol", str(stock.symbol)),
            ("Last Traded Price (LTP)", f"₹ {stock.ltp:,.2f}" if stock.ltp else "-", "Day Open", f"₹ {stock.open:,.2f}" if stock.open else "-"),
            ("Change", f"₹ {stock.change:+,.2f}" if stock.change is not None else "-", "Day High", f"₹ {stock.high:,.2f}" if stock.high else "-"),
            ("% Change", f"{stock.pct_change:+.2f}%" if stock.pct_change is not None else "-", "Day Low", f"₹ {stock.low:,.2f}" if stock.low else "-"),
            ("Previous Close", f"₹ {stock.previous_close:,.2f}" if stock.previous_close else "-", "VWAP / Average Price", f"₹ {float(vwap_val):,.2f}" if vwap_val else "-"),
            ("52-Week High", f"₹ {stock.year_high:,.2f}" if stock.year_high else "-", "52-Week Low", f"₹ {stock.year_low:,.2f}" if stock.year_low else "-"),
            ("52W High Date", yr_high_dt, "52W Low Date", yr_low_dt),
            ("30-Day % Change", f"{stock.per_change_30d:+.2f}%" if stock.per_change_30d is not None else "-", "365-Day % Change", f"{stock.per_change_365d:+.2f}%" if stock.per_change_365d is not None else "-"),
            ("Distance to 52W High", f"{stock.near_wkh:+.2f}%" if stock.near_wkh is not None else "-", "Distance to 52W Low", f"{stock.near_wkl:+.2f}%" if stock.near_wkl is not None else "-"),
            ("Price Band / Circuit", price_band, "Tick Size", tick_size),
        ]

        for lbl1, v1, lbl2, v2 in price_rows:
            c_lbl1 = ws_stock.cell(cur_row, 1, lbl1)
            c_v1 = ws_stock.cell(cur_row, 2, v1)
            c_lbl2 = ws_stock.cell(cur_row, 3, lbl2)
            c_v2 = ws_stock.cell(cur_row, 4, v2)
            format_cell(c_lbl1, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v1, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            format_cell(c_lbl2, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v2, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

        cur_row += 1

        # ----------------------------------------------------
        # SECTION 2: TRADING, VOLUME & DELIVERY ANALYTICS
        # ----------------------------------------------------
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec2 = ws_stock.cell(cur_row, 1, "2. Trading, Volume & Delivery Analytics")
        format_cell(sec2, font=HEADER_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 24
        cur_row += 1

        deliv_pct = (detail.delivery_pct if detail and detail.delivery_pct is not None else None)
        if deliv_pct is None:
            deliv_pct = t_info.get("deliveryToTradedQuantity") or s_info.get("deliveryTotradedQuantity") or t_info.get("deliveryQuantity")

        deliv_qty = t_info.get("deliveryquantity") or s_info.get("deliveryQuantity")
        daily_vol = (detail.daily_volatility if detail and detail.daily_volatility is not None else None) or p_info.get("cmDailyVolatility") or p_info.get("dailyVolatility")
        annual_vol = (detail.annual_volatility if detail and detail.annual_volatility is not None else None) or p_info.get("cmAnnualVolatility") or p_info.get("annualisedVolatility")
        margin_val = (detail.applicable_margin if detail and detail.applicable_margin is not None else None) or t_info.get("applicableMargin") or p_info.get("applicableMargin")
        var_margin = s_info.get("varMargin") or s_info.get("securityvar")
        elm_margin = s_info.get("extremelossMargin")
        sec_var = s_info.get("securityvar")
        impact_cost = (detail.impact_cost if detail and detail.impact_cost is not None else None) or t_info.get("impactCost")
        face_val = (detail.face_value if detail and detail.face_value is not None else None) or t_info.get("faceValue") or s_info.get("faceValue")
        total_mcap = t_info.get("totalMarketCap")
        total_mcap_cr = (float(total_mcap) / 10000000.0) if total_mcap else None

        ffmc_cr = (detail.free_float_mcap / 10000000.0) if (detail and detail.free_float_mcap) else ((stock.ffmc / 10000000.0) if stock.ffmc else None)
        vol_shares = stock.volume or (detail.total_volume if detail else None) or t_info.get("totalTradedVolume") or t_info.get("quantitytraded")
        turnover_cr = (detail.total_turnover / 10000000.0) if (detail and detail.total_turnover) else ((stock.turnover / 10000000.0) if stock.turnover else None)

        trade_rows = [
            ("Date", str(stock.date or target_date), "Total Turnover (₹ Cr)", f"₹ {turnover_cr:,.2f} Cr" if turnover_cr is not None else "-"),
            ("Total Traded Volume (Shares)", f"{int(vol_shares):,}" if vol_shares is not None else "-", "Delivery Quantity (Shares)", f"{int(deliv_qty):,}" if deliv_qty else "-"),
            ("Delivery %", f"{float(deliv_pct):.2f}%" if deliv_pct is not None else "-", "Free Float MCap (₹ Cr)", f"₹ {ffmc_cr:,.2f} Cr" if ffmc_cr is not None else "-"),
            ("Total Market Cap (₹ Cr)", f"₹ {total_mcap_cr:,.2f} Cr" if total_mcap_cr is not None else "-", "Impact Cost (%)", f"{float(impact_cost):.2f}%" if impact_cost is not None else "-"),
            ("Daily Volatility (%)", f"{float(daily_vol):.2f}%" if daily_vol is not None else "-", "Annualised Volatility (%)", f"{float(annual_vol):.2f}%" if annual_vol is not None else "-"),
            ("Applicable Margin (%)", f"{float(margin_val):.2f}%" if margin_val is not None else "-", "VAR Margin (%)", f"{float(var_margin):.2f}%" if var_margin else "-"),
            ("Extreme Loss Margin (%)", f"{float(elm_margin):.2f}%" if elm_margin else "-", "Security VAR (%)", f"{float(sec_var):.2f}%" if sec_var else "-"),
            ("Face Value (₹)", f"₹ {float(face_val):,.2f}" if face_val is not None else "-", "Market Lot", str(t_info.get("marketLot") or "1")),
        ]

        for lbl1, v1, lbl2, v2 in trade_rows:
            c_lbl1 = ws_stock.cell(cur_row, 1, lbl1)
            c_v1 = ws_stock.cell(cur_row, 2, v1)
            c_lbl2 = ws_stock.cell(cur_row, 3, lbl2)
            c_v2 = ws_stock.cell(cur_row, 4, v2)
            format_cell(c_lbl1, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v1, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            format_cell(c_lbl2, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v2, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

        cur_row += 1

        # ----------------------------------------------------
        # SECTION 3: SECURITY MASTER & CORPORATE INFORMATION
        # ----------------------------------------------------
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec3 = ws_stock.cell(cur_row, 1, "3. Security Master, Sector Classification & Corporate Info")
        format_cell(sec3, font=HEADER_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 24
        cur_row += 1

        isin_val = (detail.isin if detail and detail.isin else None) or m_data.get("isinCode") or s_info.get("isin") or m_data.get("isin")
        ind_val = (detail.industry if detail and detail.industry else None) or s_info.get("basicIndustry") or m_data.get("industry")
        macro_val = s_info.get("macro") or "-"
        sector_val = s_info.get("sector") or "-"
        ind_info_val = s_info.get("industryInfo") or "-"
        sector_pe = s_info.get("pdSectorPe") or "-"
        symbol_pe = s_info.get("pdSymbolPe") or "-"
        issued_cap = (detail.issued_capital if detail and detail.issued_capital is not None else None) or s_info.get("issuedSize") or t_info.get("issuedSize")
        is_fno = str(m_data.get("isFNOSec") or (stock.series == "EQ")).capitalize()
        listing_date = str(s_info.get("listingDate") or "-")
        listing_status = str(s_info.get("secStatus") or s_info.get("status") or "Active")
        board_status = str(s_info.get("boardStatus") or "Main")
        trading_segment = str(s_info.get("tradingSegment") or "Normal Market")
        class_share = str(s_info.get("classShare") or "Equity")

        sec_rows = [
            ("Date", str(stock.date or target_date), "Series", str(stock.series or "EQ")),
            ("ISIN Code", str(isin_val or "-"), "Derivatives / F&O Eligible", is_fno),
            ("Macro Sector", str(macro_val), "Sector", str(sector_val)),
            ("Basic Industry", str(ind_val or "-"), "Industry Info", str(ind_info_val)),
            ("Sector P/E", str(sector_pe), "Symbol P/E", str(symbol_pe)),
            ("Listing Date", listing_date, "Listing Status", listing_status),
            ("Board Status", board_status, "Trading Segment", trading_segment),
            ("Issued Capital (Shares)", f"{int(float(issued_cap)):,}" if issued_cap is not None else "-", "Class of Share", class_share),
        ]

        for lbl1, v1, lbl2, v2 in sec_rows:
            c_lbl1 = ws_stock.cell(cur_row, 1, lbl1)
            c_v1 = ws_stock.cell(cur_row, 2, v1)
            c_lbl2 = ws_stock.cell(cur_row, 3, lbl2)
            c_v2 = ws_stock.cell(cur_row, 4, v2)
            format_cell(c_lbl1, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v1, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            format_cell(c_lbl2, font=BOLD_FONT, align=Alignment(horizontal="left"))
            format_cell(c_v2, font=REGULAR_FONT, align=Alignment(horizontal="right"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

        cur_row += 1

        # ----------------------------------------------------
        # SECTION 4: CORPORATE ACTIONS & FINANCIAL CALENDAR
        # ----------------------------------------------------
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec4 = ws_stock.cell(cur_row, 1, "4. Corporate Actions, Announcements & Financial Calendar")
        format_cell(sec4, font=HEADER_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 24
        cur_row += 1

        stock_ca_all = actions_by_symbol.get(sym, [])

        # --- 4A. Corporate Actions & Dividends (In Chronological Order) ---
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec4a = ws_stock.cell(cur_row, 1, "4A. Corporate Actions & Dividends Timeline (Chronological Order)")
        format_cell(sec4a, font=SUBSECTION_FONT, fill=SUBSECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 20
        cur_row += 1

        ca_headers = ["Ex-Date", "Action Type", "Subject / Description", "Record Date"]
        for col_i, ch in enumerate(ca_headers, 1):
            c_ch = ws_stock.cell(cur_row, col_i, ch)
            format_cell(c_ch, font=BOLD_FONT, fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"), align=Alignment(horizontal="center"))
        ws_stock.row_dimensions[cur_row].height = 20
        cur_row += 1

        ca_items = [a for a in stock_ca_all if a.action_type in ("DIVIDEND", "SPLIT", "BONUS", "BUYBACK", "RIGHTS")]
        ca_items.sort(key=lambda x: parse_nse_date(x.ex_date), reverse=True)

        if ca_items:
            for ca_item in ca_items:
                c1 = ws_stock.cell(cur_row, 1, ca_item.ex_date or "-")
                c2 = ws_stock.cell(cur_row, 2, ca_item.action_type or "-")
                c3 = ws_stock.cell(cur_row, 3, ca_item.subject or "-")
                c4 = ws_stock.cell(cur_row, 4, ca_item.record_date or "-")

                badge_fill = DIVIDEND_FILL if ca_item.action_type == "DIVIDEND" else (
                    SPLIT_FILL if ca_item.action_type in ("SPLIT", "BONUS") else OTHER_FILL
                )

                format_cell(c1, font=REGULAR_FONT, align=Alignment(horizontal="center"))
                format_cell(c2, font=BOLD_FONT, fill=badge_fill, align=Alignment(horizontal="center"))
                format_cell(c3, font=REGULAR_FONT, align=Alignment(horizontal="left"))
                format_cell(c4, font=REGULAR_FONT, align=Alignment(horizontal="center"))
                ws_stock.row_dimensions[cur_row].height = 20
                cur_row += 1
        else:
            for col_i in range(1, 5):
                c = ws_stock.cell(cur_row, col_i, "-")
                format_cell(c, font=REGULAR_FONT, align=Alignment(horizontal="center"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

        cur_row += 1

        # --- 4B. Financial Calendar & Board Meetings (In Chronological Order) ---
        ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
        sec4b = ws_stock.cell(cur_row, 1, "4B. Financial Calendar & Board Meetings (Chronological Order)")
        format_cell(sec4b, font=SUBSECTION_FONT, fill=SUBSECTION_FILL, align=Alignment(horizontal="left"))
        ws_stock.row_dimensions[cur_row].height = 20
        cur_row += 1

        bm_headers = ["Event Date", "Purpose / Agenda", "Details / Description", "Category"]
        for col_i, ch in enumerate(bm_headers, 1):
            c_ch = ws_stock.cell(cur_row, col_i, ch)
            format_cell(c_ch, font=BOLD_FONT, fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"), align=Alignment(horizontal="center"))
        ws_stock.row_dimensions[cur_row].height = 20
        cur_row += 1

        bm_items = [a for a in stock_ca_all if a.action_type in ("RESULTS", "BOARD_MEETING", "AGM", "OTHER")]
        bm_items.sort(key=lambda x: parse_nse_date(x.ex_date), reverse=True)

        if bm_items:
            for bm_item in bm_items:
                c1 = ws_stock.cell(cur_row, 1, bm_item.ex_date or "-")
                c2 = ws_stock.cell(cur_row, 2, bm_item.subject or "-")
                c3 = ws_stock.cell(cur_row, 3, bm_item.details or bm_item.subject or "-")
                c4 = ws_stock.cell(cur_row, 4, bm_item.action_type or "BOARD_MEETING")

                format_cell(c1, font=REGULAR_FONT, align=Alignment(horizontal="center"))
                format_cell(c2, font=BOLD_FONT, fill=RESULTS_FILL, align=Alignment(horizontal="left"))
                format_cell(c3, font=REGULAR_FONT, align=Alignment(horizontal="left"))
                format_cell(c4, font=REGULAR_FONT, align=Alignment(horizontal="center"))
                ws_stock.row_dimensions[cur_row].height = 20
                cur_row += 1
        else:
            for col_i in range(1, 5):
                c = ws_stock.cell(cur_row, col_i, "-")
                format_cell(c, font=REGULAR_FONT, align=Alignment(horizontal="center"))
            ws_stock.row_dimensions[cur_row].height = 20
            cur_row += 1

        cur_row += 1

        # ----------------------------------------------------
        # SECTION 5: BENCHMARK & INDEX PARTICIPATION
        # ----------------------------------------------------
        index_list = s_info.get("indexList") or []
        if isinstance(index_list, list) and index_list:
            ws_stock.merge_cells(f"A{cur_row}:D{cur_row}")
            sec5 = ws_stock.cell(cur_row, 1, "5. Benchmark & Index Memberships")
            format_cell(sec5, font=HEADER_FONT, fill=SECTION_FILL, align=Alignment(horizontal="left"))
            ws_stock.row_dimensions[cur_row].height = 24
            cur_row += 1

            # Render indices in a 2-column grid
            for idx_i in range(0, len(index_list), 2):
                idx_a = index_list[idx_i]
                idx_b = index_list[idx_i + 1] if idx_i + 1 < len(index_list) else ""

                ws_stock.merge_cells(f"A{cur_row}:B{cur_row}")
                c_a = ws_stock.cell(cur_row, 1, f"• {idx_a}")
                format_cell(c_a, font=REGULAR_FONT, fill=INDEX_TAG_FILL, align=Alignment(horizontal="left"))

                if idx_b:
                    ws_stock.merge_cells(f"C{cur_row}:D{cur_row}")
                    c_b = ws_stock.cell(cur_row, 3, f"• {idx_b}")
                    format_cell(c_b, font=REGULAR_FONT, fill=INDEX_TAG_FILL, align=Alignment(horizontal="left"))
                else:
                    ws_stock.merge_cells(f"C{cur_row}:D{cur_row}")
                    c_empty = ws_stock.cell(cur_row, 3, "")
                    format_cell(c_empty, fill=INDEX_TAG_FILL)

                ws_stock.row_dimensions[cur_row].height = 18
                cur_row += 1

        ws_stock.freeze_panes = "A2"
        auto_fit_columns(ws_stock, max_col_width=52)

    wb.save(output_path)
    return output_path

async def build_broad_market_workbook(target_date: str, output_path: str) -> str:
    """Generates broad_market_indices_YYYY-MM-DD.xlsx with complete metrics across categorized sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    categories = [
        ("Broad Market", "Broad Market", "4F46E5"),
        ("Sectoral", "Sectoral", "0284C7"),
        ("Thematic", "Thematic", "D97706"),
        ("Strategy", "Strategy", "7C3AED"),
    ]

    async with AsyncSessionLocal() as db:
        q_test = await db.execute(
            select(IndexDaily).where(IndexDaily.date == target_date).limit(1)
        )
        if not q_test.scalars().first():
            q_latest = await db.execute(
                select(IndexDaily.date).order_by(IndexDaily.date.desc()).limit(1)
            )
            latest_d = q_latest.scalars().first()
            if latest_d:
                target_date = latest_d

        for cat_name, cat_query, tab_color in categories:
            ws = wb.create_sheet(title=cat_name)
            ws.sheet_properties.tabColor = tab_color

            q = await db.execute(
                select(IndexDaily).where(
                    IndexDaily.date == target_date,
                    IndexDaily.index_category == cat_query
                ).order_by(IndexDaily.pct_change.desc())
            )
            indices: List[IndexDaily] = q.scalars().all()

            # Complete Index Headers
            headers = [
                "Date", "Index Name", "Index Symbol", "Current Value", "Variation", "% Change",
                "Open", "High", "Low", "Prev Close", "P/E", "P/B", "Div Yield (%)",
                "Advances", "Declines", "Unchanged", "30D % Change", "365D % Change",
                "52W High", "52W Low", "1-Week Return (%)", "1-Month Return (%)", "1-Year Return (%)",
                "1-Week Ago Val", "1-Month Ago Val", "1-Year Ago Val"
            ]

            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                format_cell(
                    cell,
                    font=HEADER_FONT,
                    fill=HEADER_FILL,
                    align=Alignment(horizontal="center", vertical="center", wrap_text=True)
                )
            ws.row_dimensions[1].height = 28

            row_start = 2
            for idx in indices:
                pct_dec = (idx.pct_change / 100.0) if idx.pct_change is not None else None
                dy_dec = (idx.dy / 100.0) if idx.dy is not None else None
                p30_dec = (idx.per_change_30d / 100.0) if idx.per_change_30d is not None else None
                p365_dec = (idx.per_change_365d / 100.0) if idx.per_change_365d is not None else None

                # Return calculations against historical baselines
                val = idx.value or 0
                ret_1w = ((val - idx.one_week_ago_val) / idx.one_week_ago_val) if (idx.one_week_ago_val and idx.one_week_ago_val > 0 and val > 0) else None
                ret_1m = ((val - idx.one_month_ago_val) / idx.one_month_ago_val) if (idx.one_month_ago_val and idx.one_month_ago_val > 0 and val > 0) else None
                ret_1y = ((val - idx.one_year_ago_val) / idx.one_year_ago_val) if (idx.one_year_ago_val and idx.one_year_ago_val > 0 and val > 0) else None

                row_data = [
                    idx.date or target_date,
                    idx.index_name,
                    idx.index_symbol or idx.index_name,
                    idx.value,
                    idx.variation,
                    pct_dec,
                    idx.open,
                    idx.high,
                    idx.low,
                    idx.previous_close,
                    idx.pe,
                    idx.pb,
                    dy_dec,
                    idx.advances,
                    idx.declines,
                    idx.unchanged,
                    p30_dec,
                    p365_dec,
                    idx.year_high,
                    idx.year_low,
                    ret_1w,
                    ret_1m,
                    ret_1y,
                    idx.one_week_ago_val,
                    idx.one_month_ago_val,
                    idx.one_year_ago_val
                ]
                ws.append(row_data)

            row_end = len(indices) + 1

            for r in range(row_start, row_end + 1):
                is_zebra = (r % 2 == 1)
                row_fill = ZEBRA_FILL if is_zebra else None

                format_cell(ws.cell(r, 1), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="center"))
                format_cell(ws.cell(r, 2), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="left"))
                format_cell(ws.cell(r, 3), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="left"))
                format_cell(ws.cell(r, 4), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 5), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+#,##0.00;-#,##0.00;0.00")
                format_cell(ws.cell(r, 6), font=BOLD_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 7), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 8), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 9), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 10), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 11), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="0.00")
                format_cell(ws.cell(r, 12), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="0.00")
                format_cell(ws.cell(r, 13), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="0.00%")
                format_cell(ws.cell(r, 14), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0")
                format_cell(ws.cell(r, 15), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0")
                format_cell(ws.cell(r, 16), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0")
                format_cell(ws.cell(r, 17), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 18), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 19), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 20), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 21), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 22), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 23), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="+0.00%;-0.00%;0.00%")
                format_cell(ws.cell(r, 24), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 25), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                format_cell(ws.cell(r, 26), font=REGULAR_FONT, fill=row_fill, align=Alignment(horizontal="right"), num_format="#,##0.00")
                ws.row_dimensions[r].height = 20

            if row_end >= row_start:
                color_scale = ColorScaleRule(
                    start_type="num", start_value=-0.03, start_color="FCA5A5",
                    mid_type="num", mid_value=0.0, mid_color="FFFFFF",
                    end_type="num", end_value=0.03, end_color="86EFAC"
                )
                ws.conditional_formatting.add(f"F2:F{row_end}", color_scale)
                ws.conditional_formatting.add(f"Q2:R{row_end}", color_scale)
                ws.conditional_formatting.add(f"U2:W{row_end}", color_scale)

                icon_set = IconSetRule("3Arrows", "num", [0, 0.0001], showValue=None, reverse=None)
                ws.conditional_formatting.add(f"F2:F{row_end}", icon_set)

            ws.freeze_panes = "C2"
            ws.auto_filter.ref = f"A1:Z{max(row_end, 1)}"
            auto_fit_columns(ws)

    wb.save(output_path)
    return output_path

async def generate_full_export_bundle(target_date: Optional[str] = None) -> Tuple[str, List[str], str]:
    """Builds both workbooks in a dated directory and packages them into a single ZIP.
    If executed on a weekend or government holiday, falls back to the latest actual synced market trade date from NSE.
    """
    fetcher = NSEFetcher()
    async with AsyncSessionLocal() as db:
        resolved_date = target_date
        if resolved_date:
            try:
                dt_req = datetime.strptime(resolved_date, "%Y-%m-%d").date()
                is_closed, _ = fetcher.is_market_holiday_or_weekend(dt_req)
                if is_closed:
                    resolved_date = None
                else:
                    q_chk = await db.execute(select(Nifty50Daily.date).where(Nifty50Daily.date == resolved_date).limit(1))
                    if not q_chk.scalars().first():
                        resolved_date = None
            except Exception:
                resolved_date = None

        if not resolved_date:
            q_dates = await db.execute(select(Nifty50Daily.date).distinct().order_by(desc(Nifty50Daily.date)))
            all_dates = q_dates.scalars().all()
            for d_str in all_dates:
                try:
                    d_obj = datetime.strptime(d_str, "%Y-%m-%d").date()
                    is_hol, _ = fetcher.is_market_holiday_or_weekend(d_obj)
                    if not is_hol:
                        resolved_date = d_str
                        break
                except Exception:
                    pass
            if not resolved_date and all_dates:
                resolved_date = all_dates[0]

        if not resolved_date:
            resolved_date = date.today().strftime("%Y-%m-%d")

    target_date = resolved_date

    export_dir = Path(settings.EXPORT_DIR) / target_date
    export_dir.mkdir(parents=True, exist_ok=True)

    nifty_file = export_dir / f"nifty50_daily_{target_date}.xlsx"
    indices_file = export_dir / f"broad_market_indices_{target_date}.xlsx"
    zip_file = export_dir / f"NSE_Market_Data_Export_{target_date}.zip"

    await build_nifty50_workbook(target_date, str(nifty_file))
    await build_broad_market_workbook(target_date, str(indices_file))

    # Synchronize and include Master Multi-Sheet Workbooks (all historical dates appended)
    from app.services.excel_sync import master_excel_sync
    idx_master_path, n50_master_path = await master_excel_sync.sync_all_masters()

    with zipfile.ZipFile(zip_file, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(nifty_file, arcname=nifty_file.name)
        z.write(indices_file, arcname=indices_file.name)
        if os.path.exists(n50_master_path):
            z.write(n50_master_path, arcname="nifty50_daily_master.xlsx")
        if os.path.exists(idx_master_path):
            z.write(idx_master_path, arcname="broad_market_indices_master.xlsx")

    files = [str(nifty_file), str(indices_file), str(n50_master_path), str(idx_master_path)]
    return str(zip_file), files, target_date
