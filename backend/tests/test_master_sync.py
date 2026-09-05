import os
import pytest
import openpyxl
from datetime import date
from app.services.excel_sync import MasterExcelSyncManager
from app.services.db_manager import DatabaseManager
from app.services.nse_scraper import NSEScraper

@pytest.mark.asyncio
async def test_nse_scraper_session():
    scraper = NSEScraper()
    assert scraper.BASE_URL == "https://www.nseindia.com"
    is_open, status_text = scraper.is_market_open()
    assert isinstance(is_open, bool)
    assert isinstance(status_text, str)

@pytest.mark.asyncio
async def test_database_manager_queries():
    dates = await DatabaseManager.get_available_trade_dates()
    assert isinstance(dates, list)
    
    indices = await DatabaseManager.get_all_indices_history()
    assert isinstance(indices, list)
    
    stocks = await DatabaseManager.get_all_nifty50_history()
    assert isinstance(stocks, list)

@pytest.mark.asyncio
async def test_master_workbooks_generation(tmp_path):
    sync_mgr = MasterExcelSyncManager(output_dir=tmp_path)
    idx_path, n50_path = await sync_mgr.sync_all_masters()
    
    assert os.path.exists(idx_path)
    assert os.path.exists(n50_path)
    
    # Verify Indices Master sheets
    wb_idx = openpyxl.load_workbook(idx_path, read_only=True)
    assert "Broad Market" in wb_idx.sheetnames
    assert "Sectoral" in wb_idx.sheetnames
    assert "Thematic" in wb_idx.sheetnames
    assert "Strategy" in wb_idx.sheetnames
    wb_idx.close()
    
    # Verify Nifty 50 Master sheets
    wb_n50 = openpyxl.load_workbook(n50_path, read_only=True)
    assert "Nifty 50 Overview" in wb_n50.sheetnames
    wb_n50.close()
